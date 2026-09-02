"""Tennis backtest — the client's brief, slams 2020-2026, ATP + WTA.

The question: a player priced 60c+ before the match LOSES THE FIRST SET.
How often do they still win — and what did it pay to back them after the
set-one loss?

Sources, joined per match:
  * tennis-data.co.uk  - one Excel per tour per season: every completed
    match with SET-BY-SET scores and pre-match bookmaker odds. The odds,
    de-vigged, define "60c+ favorite" for every year — including the years
    before Polymarket listed tennis at all.
  * Polymarket Gamma/CLOB (2024 on) - the real match markets under the slam
    tags. 2025-26 coverage is broad; 2024 lists late rounds only.

"Price after set one": no free source stamps WHEN a set ended, so the
favorite's price is sampled at match start + 45 minutes — the median length
of a first set — window 35-60', nearest 45'. An approximation, labeled as
such, mirroring football.py's kickoff+75' convention.

Slams per the client's spec: Australian Open, Roland Garros, Wimbledon.
Research CLI, runs anywhere (only public endpoints, no keys):
    python -m backend.backtest.tennis
Every download and price pull is cached in CACHE for cheap reruns.
"""
import io
import json
import os
import re
import tempfile
import time
import unicodedata
from datetime import datetime, timedelta, timezone

import httpx

CACHE = os.path.join(tempfile.gettempdir(), "tennis_backtest_cache.json")
RESULTS = os.path.join(tempfile.gettempdir(), "tennis_backtest_results.json")
XLSX_DIR = os.path.join(tempfile.gettempdir(), "tennis_xlsx")

YEARS = list(range(2020, 2027))
FAV_MIN = 0.60                       # the "60c+" bar, de-vigged
SLAMS = {                            # tennis-data Tournament -> gamma tags
    "Australian Open": ["australian-open"],
    "French Open": ["french-open"],
    "Wimbledon": ["wimbledon"],
}
PRICE_OFFSET_MIN = 45                # ~median first-set length
PRICE_WINDOW = (35, 60)


def norm(s: str) -> str:
    s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode()
    return re.sub(r"\s+", " ", re.sub(r"[^a-z ]", " ", s.lower())).strip()


def surname(td_name: str) -> str:
    """tennis-data names look like 'Alcaraz C.' or 'Van De Zandschulp B.' —
    everything except trailing single-letter initials is the surname."""
    toks = [t for t in norm(td_name).split() if len(t) > 1]
    return " ".join(toks)


def _load_cache() -> dict:
    try:
        return json.load(open(CACHE))
    except Exception:  # noqa: BLE001
        return {}


def _save_cache(c: dict):
    json.dump(c, open(CACHE, "w"))


# ------------------------------------------------------- tennis-data files
def fetch_sheets() -> list[dict]:
    """Every slam match 2020-2026, both tours, as plain dicts."""
    from openpyxl import load_workbook
    os.makedirs(XLSX_DIR, exist_ok=True)
    http = httpx.Client(timeout=60, follow_redirects=True,
                        headers={"User-Agent": "Mozilla/5.0"})
    rows = []
    for year in YEARS:
        for tour, candidates in (
                ("ATP", [f"{year}/{year}.xlsx"]),
                ("WTA", [f"{year}w/{year}w.xlsx", f"{year}w/{year}.xlsx"])):
            path = os.path.join(XLSX_DIR, f"{tour}-{year}.xlsx")
            if not os.path.exists(path):
                for cand in candidates:
                    try:
                        r = http.get("http://www.tennis-data.co.uk/" + cand)
                        if r.status_code == 200 and len(r.content) > 10000:
                            open(path, "wb").write(r.content)
                            break
                    except Exception:  # noqa: BLE001
                        pass
            if not os.path.exists(path):
                print("  MISSING %s %d (no file)" % (tour, year))
                continue
            wb = load_workbook(io.BytesIO(open(path, "rb").read()), read_only=True)
            ws = wb[wb.sheetnames[0]]
            header = None
            n = 0
            for raw in ws.iter_rows(values_only=True):
                if header is None:
                    header = [str(x or "") for x in raw]
                    continue
                d = dict(zip(header, raw))
                if str(d.get("Tournament") or "") not in SLAMS:
                    continue
                d["_tour"], d["_year"] = tour, year
                rows.append(d)
                n += 1
            print("  %s %d: %d slam rows" % (tour, year, n))
    return rows


def _num(v):
    try:
        f = float(v)
        return f if f == f else None                  # NaN guard
    except (TypeError, ValueError):
        return None


def qualify(rows: list[dict]) -> list[dict]:
    """Completed matches with a de-vigged >=60% favorite who LOST set one."""
    out, skipped_odds = [], 0
    for d in rows:
        if str(d.get("Comment") or "").strip().lower() != "completed":
            continue
        w1, l1 = _num(d.get("W1")), _num(d.get("L1"))
        if w1 is None or l1 is None or w1 == l1:
            continue
        # market odds: Avg preferred, then Pinnacle, then B365
        for a, b in (("AvgW", "AvgL"), ("PSW", "PSL"), ("B365W", "B365L")):
            ow, ol = _num(d.get(a)), _num(d.get(b))
            if ow and ol and ow > 1 and ol > 1:
                break
        else:
            skipped_odds += 1
            continue
        pw = (1 / ow) / (1 / ow + 1 / ol)              # de-vigged P(match winner)
        fav_is_winner = pw >= 0.5
        fav_p = pw if fav_is_winner else 1 - pw
        if fav_p < FAV_MIN:
            continue
        fav_lost_s1 = (w1 < l1) if fav_is_winner else (l1 < w1)
        if not fav_lost_s1:
            continue
        fav, opp = (d["Winner"], d["Loser"]) if fav_is_winner else (d["Loser"], d["Winner"])
        date = d.get("Date")
        out.append({
            "tour": d["_tour"], "year": d["_year"],
            "slam": str(d["Tournament"]), "round": str(d.get("Round") or ""),
            "date": date.strftime("%Y-%m-%d") if hasattr(date, "strftime") else str(date)[:10],
            "fav": str(fav), "opp": str(opp),
            "fav_p": round(fav_p, 3), "fav_won": fav_is_winner,
            "score_s1": "%d-%d" % ((w1, l1) if fav_is_winner else (l1, w1)),
        })
    if skipped_odds:
        print("  (%d slam matches had no odds columns)" % skipped_odds)
    return out


# ---------------------------------------------------------- polymarket side
def fetch_slam_events(cache: dict) -> list[dict]:
    ck = "events"
    if ck in cache:
        return cache[ck]
    g = httpx.Client(base_url="https://gamma-api.polymarket.com", timeout=30)
    events, seen = [], set()
    tags = sorted({t for lst in SLAMS.values() for t in lst} | {"tennis"})
    for slug in tags:
        r = g.get("/tags/slug/" + slug)
        if r.status_code != 200 or not r.json().get("id"):
            continue
        tid, off = int(r.json()["id"]), 0
        while True:
            r2 = g.get("/events", params={
                "tag_id": tid, "closed": "true", "limit": 100, "offset": off,
                "start_date_min": "2024-01-01T00:00:00Z"})
            if r2.status_code != 200 or not r2.json():
                break
            for e in r2.json():
                if e.get("id") in seen:
                    continue
                seen.add(e.get("id"))
                title = e.get("title") or ""
                if not re.search(r"\svs\.?\s", title, re.I):
                    continue
                markets = []
                for m in e.get("markets") or []:
                    markets.append({
                        "question": m.get("question"),
                        "groupItemTitle": m.get("groupItemTitle"),
                        "outcomes": m.get("outcomes"),
                        "clobTokenIds": m.get("clobTokenIds"),
                        "gameStartTime": m.get("gameStartTime"),
                    })
                events.append({"slug": e.get("slug"), "title": title,
                               "start": e.get("startDate"), "tag": slug,
                               "markets": markets})
            off += 100
            if off > 6000:
                break
            time.sleep(0.15)
    print("  polymarket closed tennis vs-events since 2024:", len(events))
    cache[ck] = events
    _save_cache(cache)
    return events


def match_event(q: dict, events: list[dict]):
    """Find the Polymarket event for one qualifying match: both surnames in
    the title, listing within the slam window (+/-40 days of the match)."""
    f, o = surname(q["fav"]), surname(q["opp"])
    if not f or not o:
        return None
    best = None
    md = datetime.strptime(q["date"], "%Y-%m-%d")
    for e in events:
        t = norm(e["title"])
        if f not in t or o not in t:
            continue
        sd = (e.get("start") or "")[:10]
        try:
            gap = abs((datetime.strptime(sd, "%Y-%m-%d") - md).days)
        except ValueError:
            continue
        if gap > 40:
            continue
        if best is None or gap < best[0]:
            best = (gap, e)
    return best[1] if best else None


def fav_price_series(cache: dict, ev: dict, fav_surname: str, clob) -> dict | None:
    """Pre-match price and the +45' sample for the favorite's win market."""
    key = "px:%s:%s" % (ev["slug"], fav_surname)
    if key in cache:
        return cache[key]
    out = None
    try:
        token = start = None
        for m in ev["markets"]:
            label = norm("%s %s" % (m.get("question") or "", m.get("groupItemTitle") or ""))
            outcomes = json.loads(m.get("outcomes") or "[]")
            tokens = json.loads(m.get("clobTokenIds") or "[]")
            if not tokens:
                continue
            if outcomes and norm(outcomes[0]) not in ("yes", "no"):
                for i, oc in enumerate(outcomes):
                    if fav_surname in norm(oc) and i < len(tokens):
                        token, start = tokens[i], m.get("gameStartTime")
                        break
            elif fav_surname in label:
                token, start = tokens[0], m.get("gameStartTime")
            if token:
                break
        if token and start:
            k = datetime.fromisoformat(str(start).replace("Z", "+00:00"))
            lo = int((k - timedelta(minutes=30)).timestamp())
            hi = int((k + timedelta(minutes=PRICE_WINDOW[1])).timestamp())
            h = clob.get("/prices-history", params={
                "market": token, "startTs": lo, "endTs": hi,
                "fidelity": 1}).json().get("history") or []
            time.sleep(0.15)
            pre = [p for p in h if p["t"] <= k.timestamp()]
            mid = [p for p in h if k.timestamp() + PRICE_WINDOW[0] * 60
                   <= p["t"] <= k.timestamp() + PRICE_WINDOW[1] * 60]
            want = k.timestamp() + PRICE_OFFSET_MIN * 60
            out = {
                "pre": round(float(pre[-1]["p"]) * 100, 1) if pre else None,
                "at45": round(float(min(mid, key=lambda p: abs(p["t"] - want))["p"])
                              * 100, 1) if mid else None,
            }
    except Exception:  # noqa: BLE001 — a match without usable history has no price
        out = None
    cache[key] = out
    _save_cache(cache)
    return out


# ------------------------------------------------------------------ main
def run():
    cache = _load_cache()
    print("[1/3] tennis-data downloads")
    rows = fetch_sheets()
    q = qualify(rows)
    print("qualifying favorites (>=60%%, lost set 1): %d" % len(q))

    print("[2/3] polymarket slam events")
    events = fetch_slam_events(cache)
    clob = httpx.Client(base_url="https://clob.polymarket.com", timeout=30)

    print("[3/3] price sampling (2024+)")
    for m in q:
        m["poly"] = None
        if m["year"] < 2024:
            continue
        ev = match_event(m, events)
        if not ev:
            continue
        px = fav_price_series(cache, ev, surname(m["fav"]), clob)
        m["poly"] = {"slug": ev["slug"], **(px or {})} if px else {"slug": ev["slug"]}

    # ---- report -----------------------------------------------------------
    def bucket(matches):
        n = len(matches)
        w = sum(1 for m in matches if m["fav_won"])
        priced = [m for m in matches if m.get("poly") and m["poly"].get("at45")]
        pnl = 0.0
        for m in priced:
            e = m["poly"]["at45"]
            shares = 100.0 / (e / 100.0)
            fee = shares * 0.05 * (e / 100.0) * (1 - e / 100.0)
            pnl += (shares * (100 - e) / 100.0 if m["fav_won"] else -100.0) - fee
        return {"spots": n, "fav_won": w,
                "win_rate": round(100.0 * w / n, 1) if n else None,
                "avg_fav_p": round(100 * sum(m["fav_p"] for m in matches) / n, 1) if n else None,
                "priced": len(priced),
                "avg_at45": round(sum(m["poly"]["at45"] for m in priced)
                                  / len(priced), 1) if priced else None,
                "pnl100_fees": round(pnl, 2)}

    report = {"meta": {
        "fav_min": FAV_MIN, "years": [YEARS[0], YEARS[-1]],
        "slams": list(SLAMS), "price_note":
            "price sampled at match start +45min (median first-set length), "
            "window 35-60min — an approximation of 'just after set one'",
        "computed_at": datetime.now(timezone.utc).strftime("%Y-%m-%d")},
        "byTour": {}}
    for tour in ("ATP", "WTA"):
        tm = [m for m in q if m["tour"] == tour]
        entry = {"overall": bucket(tm), "bySlam": {}, "byYear": {}, "games": tm}
        for slam in SLAMS:
            entry["bySlam"][slam] = bucket([m for m in tm if m["slam"] == slam])
        for y in YEARS:
            ym = [m for m in tm if m["year"] == y]
            if ym:
                entry["byYear"][str(y)] = bucket(ym)
        report["byTour"][tour] = entry
        o = entry["overall"]
        print("%s: %d spots | fav won %s%% | priced %d | avg price@+45 %s | "
              "pnl after fees %+.0f" % (tour, o["spots"], o["win_rate"],
                                        o["priced"], o["avg_at45"],
                                        o["pnl100_fees"]))
    json.dump(report, open(RESULTS, "w"), indent=1)
    print("\nresults ->", RESULTS)
    return report


if __name__ == "__main__":
    run()
